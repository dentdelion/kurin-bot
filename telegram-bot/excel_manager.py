import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import config
import logging

logger = logging.getLogger(__name__)

class ExcelManager:
    def __init__(self):
        self.file_path = config.EXCEL_FILE_PATH
        
    def read_books(self):
        """Read all books from Excel file"""
        try:
            df = pd.read_excel(self.file_path)
            return df
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            return pd.DataFrame()
    
    def get_books_by_category(self, category, page=0):
        """Get books filtered by category with pagination"""
        df = self.read_books()
        if df.empty:
            return [], 0
            
        # Filter books by category
        if category != 'all':
            mask = df[config.EXCEL_COLUMNS['categories']].astype(str).str.contains(category, case=False, na=False)
            filtered_df = df[mask]
        else:
            filtered_df = df
            
        total_books = len(filtered_df)
        
        # Apply pagination
        start_idx = page * config.BOOKS_PER_PAGE
        end_idx = start_idx + config.BOOKS_PER_PAGE
        page_books = filtered_df.iloc[start_idx:end_idx]
        
        books = []
        for idx, row in page_books.iterrows():
            book_info = {
                'index': idx,
                'name': row[config.EXCEL_COLUMNS['name']],
                'author': row[config.EXCEL_COLUMNS['author']],
                'edition': row[config.EXCEL_COLUMNS['edition']],
                'pages': row[config.EXCEL_COLUMNS['pages']],
                'description': row[config.EXCEL_COLUMNS['description']],
                'booked_until': row[config.EXCEL_COLUMNS['booked_until']],
                'categories': row[config.EXCEL_COLUMNS['categories']],
                'in_queue_for_delivery': row[config.EXCEL_COLUMNS['in_queue_for_delivery']],
                'is_available': self._is_book_available(row)
            }
            books.append(book_info)
            
        return books, total_books
    
    def _is_book_available(self, row):
        """Check if book is available for booking"""
        booked_until = row[config.EXCEL_COLUMNS['booked_until']]
        in_queue = row[config.EXCEL_COLUMNS['in_queue_for_delivery']]
        
        # Book is available if booked_until is empty/null and not in delivery queue
        return (pd.isna(booked_until) or booked_until == '') and (pd.isna(in_queue) or str(in_queue).lower() != 'yes')
    
    def book_item(self, book_index, user_id, user_name):
        """Mark book as in queue for delivery"""
        try:
            df = self.read_books()
            df.loc[book_index, config.EXCEL_COLUMNS['in_queue_for_delivery']] = 'yes'
            
            # Save to Excel
            df.to_excel(self.file_path, index=False)
            logger.info(f"Book {book_index} booked by user {user_id} ({user_name})")
            return True
        except Exception as e:
            logger.error(f"Error booking book: {e}")
            return False
    
    def get_books_for_delivery(self):
        """Get books that need to be delivered"""
        df = self.read_books()
        if df.empty:
            return []
            
        mask = df[config.EXCEL_COLUMNS['in_queue_for_delivery']].astype(str).str.lower() == 'yes'
        delivery_books = df[mask]
        
        books = []
        for idx, row in delivery_books.iterrows():
            book_info = {
                'index': idx,
                'name': row[config.EXCEL_COLUMNS['name']],
                'author': row[config.EXCEL_COLUMNS['author']],
                'edition': row[config.EXCEL_COLUMNS['edition']]
            }
            books.append(book_info)
            
        return books
    
    def mark_as_delivered(self, book_index):
        """Mark book as delivered (put on shelf)"""
        try:
            df = self.read_books()
            df.loc[book_index, config.EXCEL_COLUMNS['in_queue_for_delivery']] = 'delivered'
            df.to_excel(self.file_path, index=False)
            logger.info(f"Book {book_index} marked as delivered")
            return True
        except Exception as e:
            logger.error(f"Error marking book as delivered: {e}")
            return False
    
    def mark_as_picked_up(self, book_index, user_id):
        """Mark book as picked up by user"""
        try:
            # Read Excel with openpyxl to handle styling
            wb = load_workbook(self.file_path)
            ws = wb.active
            
            # Find the row (book_index + 2 because of header and 0-indexing)
            row_num = book_index + 2
            
            # Set values
            booked_until_col = self._get_column_index(config.EXCEL_COLUMNS['booked_until'])
            in_queue_col = self._get_column_index(config.EXCEL_COLUMNS['in_queue_for_delivery'])
            
            due_date = datetime.now() + timedelta(days=config.ALLOWED_TIME_TO_READ_THE_BOOK)
            ws.cell(row=row_num, column=booked_until_col, value=due_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=in_queue_col, value='no')
            
            # Color row yellow
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_num, column=col).fill = yellow_fill
                
            wb.save(self.file_path)
            logger.info(f"Book {book_index} marked as picked up by user {user_id}")
            return True
        except Exception as e:
            logger.error(f"Error marking book as picked up: {e}")
            return False
    
    def mark_as_returned(self, book_index):
        """Mark book as returned and clear booking"""
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            
            row_num = book_index + 2
            
            # Clear values
            booked_until_col = self._get_column_index(config.EXCEL_COLUMNS['booked_until'])
            in_queue_col = self._get_column_index(config.EXCEL_COLUMNS['in_queue_for_delivery'])
            
            ws.cell(row=row_num, column=booked_until_col, value=None)
            ws.cell(row=row_num, column=in_queue_col, value='no')
            
            # Clear row background
            no_fill = PatternFill(fill_type=None)
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_num, column=col).fill = no_fill
                
            wb.save(self.file_path)
            logger.info(f"Book {book_index} marked as returned")
            return True
        except Exception as e:
            logger.error(f"Error marking book as returned: {e}")
            return False
    
    def _get_column_index(self, column_name):
        """Get column index by name"""
        df = self.read_books()
        return df.columns.get_loc(column_name) + 1  # +1 for Excel 1-indexing
    
    def get_overdue_books(self):
        """Get books that are overdue"""
        df = self.read_books()
        if df.empty:
            return []
            
        current_date = datetime.now().date()
        overdue_books = []
        
        for idx, row in df.iterrows():
            booked_until = row[config.EXCEL_COLUMNS['booked_until']]
            if pd.notna(booked_until) and booked_until != '':
                try:
                    due_date = pd.to_datetime(booked_until).date()
                    if due_date < current_date:
                        book_info = {
                            'index': idx,
                            'name': row[config.EXCEL_COLUMNS['name']],
                            'author': row[config.EXCEL_COLUMNS['author']],
                            'due_date': due_date,
                            'days_overdue': (current_date - due_date).days
                        }
                        overdue_books.append(book_info)
                except:
                    continue
                    
        return overdue_books
    
    def get_book_by_index(self, book_index):
        """Get specific book by index"""
        df = self.read_books()
        if df.empty or book_index >= len(df):
            return None
            
        row = df.iloc[book_index]
        return {
            'index': book_index,
            'name': row[config.EXCEL_COLUMNS['name']],
            'author': row[config.EXCEL_COLUMNS['author']],
            'edition': row[config.EXCEL_COLUMNS['edition']],
            'pages': row[config.EXCEL_COLUMNS['pages']],
            'description': row[config.EXCEL_COLUMNS['description']],
            'booked_until': row[config.EXCEL_COLUMNS['booked_until']],
            'categories': row[config.EXCEL_COLUMNS['categories']],
            'in_queue_for_delivery': row[config.EXCEL_COLUMNS['in_queue_for_delivery']],
            'is_available': self._is_book_available(row)
        } 